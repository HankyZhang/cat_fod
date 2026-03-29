# -*- coding: utf-8 -*-
"""
生成猫饭NRC配方计算器 Excel
- 3个版本各一个工作表，改总重量就能自动算出所有食材、补剂和29项NRC指标
- 1个"营养矩阵"参考表
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── 原始数据 ──────────────────────────────────────────────
INGREDIENTS = ['鸡腿肉', '猪瘦肉', '鸡心', '牛心', '鸡蛋', '鸡肝', '牛肝', '鲅鱼', '鸡皮', '鸭胸']

# 每100g生重
NUT_MATRIX = {
    '能量(kcal)':    [119,   143,   153,   112,   143,   119,   135,   139,    349,   135],
    '蛋白质(g)':     [19.6,  21.1,  15.6,  17.7,  12.6,  16.9,  20.4,  19.29,  13.3,  18.28],
    '脂肪(g)':       [4.3,   6.9,   9.3,   3.9,   9.9,   4.8,   3.6,   6.30,   32.4,  5.95],
    '磷(mg)':        [178,   197,   177,   212,   198,   297,   387,   171,     68,    203],
    '钙(mg)':        [9,     5,     12,    7,     56,    8,     5,     11,      11,    11],
    '铁(mg)':        [0.8,   0.8,   5.9,   4.3,   1.8,   9.0,   4.9,   0.44,   0.7,   2.4],
    '锌(mg)':        [1.8,   1.9,   6.6,   2.4,   1.3,   2.7,   4.0,   0.51,   0.6,   1.9],
    '铜(mg)':        [0.06,  0.06,  0.33,  0.39,  0.07,  0.49,  9.76,  0.046,  0.02,  0.24],
    '锰(mg)':        [0.02,  0.01,  0.05,  0.03,  0.03,  0.26,  0.31,  0.015,  0.01,  0.02],
    '钾(mg)':        [242,   389,   176.066, 286.984, 138, 230,  313,   446,    119,   271],
    '钠(mg)':        [95,    52,    73.934, 97.989, 142,  71,    69,    59,     51,    74],
    '碘(mcg)':       [1,     1,     5,     5,     27,    3,     3,     35,     1,     1],
    '硒(mcg)':       [23.0,  32.5,  16.0,  21.8,  30.7,  54.6,  39.7,  36.5,   14.1,  14.0],
    'VA(IU)':        [50,    7,     31,    0,     520,   11078, 16898, 98,     150,   80],
    'VD(IU)':        [0,     12,    0,     0,     82,    8,     49,    292,    24,    4],
    'VE(mg)':        [0.2,   0.5,   0.7,   0.2,   1.1,   0.7,   0.5,   0.42,   0.3,   0.7],
    'VB1(mg)':       [0.07,  0.81,  0.13,  0.24,  0.04,  0.31,  0.19,  0.068,  0.01,  0.36],
    'VB2(mg)':       [0.16,  0.23,  0.73,  0.91,  0.46,  1.78,  2.76,  0.130,  0.07,  0.47],
    'VB3(mg)':       [5.3,   5.0,   4.8,   7.4,   0.1,   9.7,   13.2,  8.47,   2.5,   5.3],
    'VB6(mg)':       [0.33,  0.42,  0.29,  0.28,  0.14,  0.85,  1.08,  0.315,  0.05,  0.34],
    'VB12(mcg)':     [0.4,   0.6,   7.3,   8.6,   0.9,   16.6,  59.3,  7.4,    0.3,   0.4],
    '牛磺酸(mg)':    [40,    50,    145,   160,   20,    30,    40,    70,     10,    0],
    '花生四烯酸(mg)': [50,    70,    84,    25,    155,   260,   200,   60,     250,   0],
    'EPA+DHA(mg)':   [0,     0,     0,     0,     0,     0,     0,     941,    0,     0],
    '亚油酸(g)':     [0.748, 0.49,  1.918, 0.395, 1.554, 0.475, 0.299, 0.100,  8.288, 0.752],
    '镁(mg)':        [23,    25,    15,    21,    12,    19,    20,    33,     10,    20],
    '胆碱(mg)':      [65,    80,    65,    150,   293,   290,   333,   65,     46,    64],
    '叶酸(mcg)':     [6,     1,     72,    3,     47,    588,   290,   2,      1,     25],
    '泛酸(mg)':      [1.01,  0.60,  2.51,  2.25,  1.53,  6.50,  7.17,  0.42,   0.53,  1.60],
}

NUT_KEYS = list(NUT_MATRIX.keys())

# 版本配置: (名称, 各食材占比%, 标准总量g)
VERSIONS = {
    'A版：7%鸡皮基础版': {
        'pcts': [28.03, 18.94, 12.94, 14.05, 8.02, 4.01, 3.00, 4.01, 7.01, 0],
        'grams': [888, 600, 410, 445, 254, 127, 95, 127, 222, 0],
        'std_total': 3168,
    },
    'B版：7%鸡皮+鸭胸版': {
        'pcts': [9.09, 18.94, 12.94, 14.05, 8.02, 4.01, 3.00, 4.01, 7.01, 18.94],
        'grams': [288, 600, 410, 445, 254, 127, 95, 127, 222, 600],
        'std_total': 3168,
    },
    'C版：鸭胸强化版': {
        'pcts': [0, 18.87, 12.89, 13.99, 7.99, 3.99, 2.99, 3.99, 6.98, 28.30],
        'grams': [0, 600, 410, 445, 254, 127, 95, 127, 222, 900],
        'std_total': 3180,
    },
}

# 标准批次补充剂（基于 ~3168-3180g 原料）
# 格式: (名称, 标准用量描述, 提供的营养素字典)
SUPP_ITEMS = [
    ('碳酸钙粉',           '18g',    {'钙(mg)': 18*0.4004*1000}),
    ('牛磺酸粉',           '3g',     {'牛磺酸(mg)': 3000}),
    ('鱼油胶囊 10粒',      '10粒',   {'EPA+DHA(mg)': 3000, '能量(kcal)': 90}),
    ('维生素D3 1000IU',    '1粒',    {'VD(IU)': 1000}),
    ('维生素E 100IU×8',    '8粒',    {'VE(mg)': 8*100*0.6711}),
    ('维生素B族 B-50',     '2片',    {'VB1(mg)': 100, 'VB2(mg)': 100, 'VB3(mg)': 100,
                                      'VB6(mg)': 100, 'VB12(mcg)': 100,
                                      '叶酸(mcg)': 800, '泛酸(mg)': 100}),
    ('海带碘 225mcg×7',    '7粒',    {'碘(mcg)': 7*225}),
    ('碘盐',               '3g',     {'碘(mcg)': 3*25, '钠(mg)': 3*0.3934*1000,
                                      '氯化物(mg)': 3*0.6066*1000}),
    ('锌片 30mg',          '1粒',    {'锌(mg)': 30}),
    ('螯合铁 36mg',        '1粒',    {'铁(mg)': 36}),
    ('锰片 10mg',          '1粒',    {'锰(mg)': 10}),
]

# NRC 校验阈值: (营养素key, NRC下限/1000kcal, NRC上限/1000kcal或None, 单位)
NRC_CHECKS = [
    ('蛋白质(g)',      50,    None,   'g'),
    ('脂肪(g)',        22.5,  None,   'g'),
    ('钙(mg)',         720,   3000,   'mg'),
    ('磷(mg)',         640,   None,   'mg'),
    ('钾(mg)',         1300,  None,   'mg'),
    ('钠(mg)',         170,   None,   'mg'),
    ('氯化物(mg)',     240,   None,   'mg'),
    ('镁(mg)',         100,   None,   'mg'),
    ('铁(mg)',         20,    None,   'mg'),
    ('锌(mg)',         18.5,  None,   'mg'),
    ('铜(mg)',         1.2,   None,   'mg'),
    ('锰(mg)',         1.2,   None,   'mg'),
    ('碘(mcg)',        350,   9000,   'mcg'),
    ('硒(mcg)',        75,    None,   'mcg'),
    ('牛磺酸(mg)',     250,   None,   'mg'),
    ('VA(IU)',         3333,  333333, 'IU'),
    ('VD(IU)',         70,    7500,   'IU'),
    ('VE(mg)',         10,    None,   'mg'),
    ('VB1(mg)',        1.4,   None,   'mg'),
    ('VB2(mg)',        1.0,   None,   'mg'),
    ('VB3(mg)',        10,    None,   'mg'),
    ('VB6(mg)',        0.625, None,   'mg'),
    ('VB12(mcg)',      5.6,   None,   'mcg'),
    ('胆碱(mg)',       637,   None,   'mg'),
    ('叶酸(mcg)',      188,   None,   'mcg'),
    ('泛酸(mg)',       1.44,  None,   'mg'),
    ('花生四烯酸(mg)', 15,    None,   'mg'),
    ('EPA+DHA(mg)',    25,    None,   'mg'),
    ('亚油酸(g)',      1.4,   None,   'g'),
]

# ── 样式 ──────────────────────────────────────────────────
TITLE_FONT    = Font(name='微软雅黑', bold=True, size=14)
HEADER_FONT   = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL   = PatternFill('solid', fgColor='4472C4')
INPUT_FILL    = PatternFill('solid', fgColor='FFF2CC')
PASS_FILL     = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL     = PatternFill('solid', fgColor='FFC7CE')
LIGHT_FILL    = PatternFill('solid', fgColor='D9E2F3')
THIN_BORDER   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))
CENTER        = Alignment(horizontal='center', vertical='center')
LEFT          = Alignment(horizontal='left', vertical='center')

def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

def style_cell(ws, row, col, border=True, align=CENTER, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    if border:
        cell.border = THIN_BORDER
    cell.alignment = align
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell

# ── 创建营养矩阵参考表 ───────────────────────────────────
def create_matrix_sheet(wb):
    ws = wb.create_sheet('营养矩阵(参考)')
    ws.sheet_properties.tabColor = '70AD47'

    ws.cell(row=1, column=1, value='营养素（每100g生重）').font = TITLE_FONT
    ws.column_dimensions['A'].width = 18
    for i in range(10):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    # 表头
    r = 3
    ws.cell(row=r, column=1, value='营养素')
    for i, name in enumerate(INGREDIENTS):
        ws.cell(row=r, column=i + 2, value=name)
    style_header(ws, r, 1, 11)

    # 数据
    for ri, (key, vals) in enumerate(NUT_MATRIX.items()):
        row = r + 1 + ri
        style_cell(ws, row, 1, align=LEFT).value = key
        for ci, v in enumerate(vals):
            style_cell(ws, row, ci + 2, fmt='0.###').value = v

    return ws

# ── 创建版本计算表 ────────────────────────────────────────
def create_version_sheet(wb, ver_name, ver_data):
    short = ver_name.split('：')[0]  # A版 / B版 / C版
    ws = wb.create_sheet(short)
    ws.sheet_properties.tabColor = '4472C4' if 'A' in short else ('ED7D31' if 'B' in short else '7030A0')

    pcts = ver_data['pcts']
    std_grams = ver_data['grams']
    std_total = ver_data['std_total']

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14

    # ── 标题 ──
    r = 1
    ws.cell(row=r, column=1, value=f'🐱 {ver_name}').font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r = 2
    ws.cell(row=r, column=1, value='基于 NRC 2006 · 鲅鱼体系 · 建议每批补1粒VD3(不补也达标)').font = Font(name='微软雅黑', size=10, color='666666')

    # ── 总量输入 ──
    r = 4
    ws.cell(row=r, column=1, value='⬇ 修改这里的总原料量').font = Font(name='微软雅黑', bold=True, size=11, color='C00000')
    r = 5
    style_cell(ws, r, 1, align=LEFT, fill=LIGHT_FILL).value = '总原料量(g)'
    input_cell = style_cell(ws, r, 2, fill=INPUT_FILL, fmt='#,##0')
    input_cell.value = std_total
    input_cell.font = Font(name='微软雅黑', bold=True, size=13)
    total_ref = f'B{r}'  # e.g. B5

    # ── 食材表 ──
    r = 7
    ws.cell(row=r, column=1, value='食材').font = Font(name='微软雅黑', bold=True)
    ws.cell(row=r, column=2, value='占比')
    ws.cell(row=r, column=3, value='标准用量(g)')
    ws.cell(row=r, column=4, value='当前用量(g)')
    style_header(ws, r, 1, 4)

    ingr_weight_cells = []  # 记录每种食材"当前用量"的单元格地址
    for i in range(10):
        row = r + 1 + i
        style_cell(ws, row, 1, align=LEFT).value = INGREDIENTS[i]
        style_cell(ws, row, 2, fmt='0.0%').value = pcts[i] / 100
        style_cell(ws, row, 3, fmt='#,##0').value = std_grams[i]
        # 当前用量 = 总量 × 占比(精确)，用 ROUND 取整
        # 精确占比 = 标准克数 / 标准总量
        exact_pct = std_grams[i] / std_total if std_total else 0
        c = style_cell(ws, row, 4, fmt='#,##0', fill=LIGHT_FILL)
        c.value = f'=ROUND({total_ref}*{exact_pct},0)'
        ingr_weight_cells.append(f'D{row}')

    sum_row = r + 11
    style_cell(ws, sum_row, 1, align=LEFT).value = '合计'
    style_cell(ws, sum_row, 1).font = Font(name='微软雅黑', bold=True)
    style_cell(ws, sum_row, 4, fmt='#,##0').value = f'=SUM(D{r+1}:D{r+10})'

    # ── 基础结果 ──
    res_start = sum_row + 2
    ws.cell(row=res_start, column=1, value='基础结果').font = Font(name='微软雅黑', bold=True, size=11)

    def res_row(row, label, formula, fmt='#,##0.0'):
        style_cell(ws, row, 1, align=LEFT).value = label
        style_cell(ws, row, 2, fmt=fmt).value = formula

    # 成品重量
    finished_row = res_start + 1
    res_row(finished_row, '成品重量(g)(90%出成率)', f'={total_ref}*0.9')

    # 总能量 = 各食材能量之和 + 鱼油热量(按比例)
    # 食材能量 = sum(weight_i / 100 * energy_per100_i)
    energy_vals = NUT_MATRIX['能量(kcal)']
    energy_parts = []
    for i in range(10):
        if energy_vals[i] != 0 and std_grams[i] != 0:
            energy_parts.append(f'{ingr_weight_cells[i]}/100*{energy_vals[i]}')
        elif energy_vals[i] != 0:
            energy_parts.append(f'{ingr_weight_cells[i]}/100*{energy_vals[i]}')
    energy_formula = '+'.join(energy_parts)
    # 鱼油热量按比例: 90 * (当前总量/标准总量)
    fish_oil_energy = f'+90*{total_ref}/{std_total}'
    kcal_row = finished_row + 1
    kcal_formula = f'={energy_formula}{fish_oil_energy}'
    res_row(kcal_row, '总能量(kcal)', kcal_formula)
    kcal_ref = f'B{kcal_row}'

    density_row = kcal_row + 1
    res_row(density_row, '能量密度(kcal/g)', f'={kcal_ref}/B{finished_row}', fmt='0.000')

    # ── 补充剂表 ──
    supp_start = density_row + 2
    ws.cell(row=supp_start, column=1, value='补充剂（随总量等比缩放）').font = Font(name='微软雅黑', bold=True, size=11)
    supp_header = supp_start + 1
    ws.cell(row=supp_header, column=1, value='补充剂')
    ws.cell(row=supp_header, column=2, value='标准用量')
    ws.cell(row=supp_header, column=3, value='当前用量')
    style_header(ws, supp_header, 1, 3)

    scale = f'{total_ref}/{std_total}'
    for si, (sname, sdesc, _) in enumerate(SUPP_ITEMS):
        row = supp_header + 1 + si
        style_cell(ws, row, 1, align=LEFT).value = sname
        style_cell(ws, row, 2, align=LEFT).value = sdesc
        # 解析标准数值
        import re
        num = re.findall(r'[\d.]+', sdesc)
        if num:
            base_val = float(num[0])
            style_cell(ws, row, 3, fmt='0.0').value = f'=ROUND({base_val}*{scale},1)'

    # ── 营养素计算 & NRC校验 ──
    nrc_start = supp_header + len(SUPP_ITEMS) + 3
    ws.cell(row=nrc_start, column=1, value='29项NRC营养校验（每1000kcal）').font = Font(name='微软雅黑', bold=True, size=11)

    nrc_header = nrc_start + 1
    for ci, h in enumerate(['营养素', '食材总量', '补充剂', '合计', '每1000kcal', 'NRC下限', 'NRC上限', '状态'], 1):
        ws.cell(row=nrc_header, column=ci, value=h)
    style_header(ws, nrc_header, 1, 8)

    # 预计算每种营养的补充剂总量
    supp_totals = {}
    for _, _, snut in SUPP_ITEMS:
        for k, v in snut.items():
            supp_totals[k] = supp_totals.get(k, 0) + v

    for ci, (nut_key, nrc_low, nrc_high, unit) in enumerate(NRC_CHECKS):
        row = nrc_header + 1 + ci

        style_cell(ws, row, 1, align=LEFT).value = nut_key

        # 食材总量 = sum(weight_i / 100 * nut_per100_i)
        if nut_key in NUT_MATRIX:
            vals = NUT_MATRIX[nut_key]
            parts = []
            for j in range(10):
                if vals[j] != 0:
                    parts.append(f'{ingr_weight_cells[j]}/100*{vals[j]}')
            if parts:
                food_formula = '=' + '+'.join(parts)
            else:
                food_formula = 0
        elif nut_key == '氯化物(mg)':
            food_formula = 0  # 食材不提供氯化物，只来自碘盐
        else:
            food_formula = 0

        food_cell = style_cell(ws, row, 2, fmt='#,##0.0')
        food_cell.value = food_formula
        food_ref = f'B{row}'

        # 补充剂提供量（按比例缩放）
        supp_val = supp_totals.get(nut_key, 0)
        supp_cell = style_cell(ws, row, 3, fmt='#,##0.0')
        if supp_val > 0:
            supp_cell.value = f'=ROUND({supp_val}*{scale},1)'
        else:
            supp_cell.value = 0
        supp_ref = f'C{row}'

        # 合计 —— 牛磺酸特殊处理（食材部分打6折）
        total_cell = style_cell(ws, row, 4, fmt='#,##0.0')
        if nut_key == '牛磺酸(mg)':
            total_cell.value = f'={food_ref}*0.4+{supp_ref}'
        else:
            total_cell.value = f'={food_ref}+{supp_ref}'
        total_ref_cell = f'D{row}'

        # 每1000kcal
        per1k_cell = style_cell(ws, row, 5, fmt='#,##0.0')
        per1k_cell.value = f'={total_ref_cell}/{kcal_ref}*1000'
        per1k_ref = f'E{row}'

        # NRC下限
        style_cell(ws, row, 6, fmt='#,##0.0').value = nrc_low

        # NRC上限
        style_cell(ws, row, 7, fmt='#,##0.0').value = nrc_high if nrc_high else '—'

        # 状态: 用IF公式
        status_cell = style_cell(ws, row, 8)
        if nrc_high:
            status_cell.value = f'=IF(AND({per1k_ref}>={nrc_low},{per1k_ref}<={nrc_high}),"✅","❌")'
        else:
            status_cell.value = f'=IF({per1k_ref}>={nrc_low},"✅","❌")'

    # ── 钙磷比 ──
    cap_row = nrc_header + 1 + len(NRC_CHECKS) + 1
    # 找钙和磷的合计单元格
    ca_total_row = None
    p_total_row = None
    for ci, (nut_key, _, _, _) in enumerate(NRC_CHECKS):
        if nut_key == '钙(mg)':
            ca_total_row = nrc_header + 1 + ci
        if nut_key == '磷(mg)':
            p_total_row = nrc_header + 1 + ci

    ws.cell(row=cap_row, column=1, value='钙磷比').font = Font(name='微软雅黑', bold=True)
    style_cell(ws, cap_row, 2, fmt='0.000').value = f'=D{ca_total_row+1}/D{p_total_row+1}'
    style_cell(ws, cap_row, 3, align=LEFT).value = '（NRC推荐 1.1–1.5）'
    cap_ref = f'B{cap_row}'
    style_cell(ws, cap_row, 5).value = f'=IF(AND({cap_ref}>=1.1,{cap_ref}<=1.5),"✅","❌")'

    # ── 通过数 ──
    pass_row = cap_row + 1
    ws.cell(row=pass_row, column=1, value='通过项数').font = Font(name='微软雅黑', bold=True)
    first_status = nrc_header + 2
    last_status = nrc_header + 1 + len(NRC_CHECKS)
    style_cell(ws, pass_row, 2).value = f'=COUNTIF(H{first_status}:H{last_status},"✅")&"/29"'

    # ── 每日分配 ──
    daily_start = pass_row + 2
    ws.cell(row=daily_start, column=1, value='每日分配（参考）').font = Font(name='微软雅黑', bold=True, size=11)
    dh = daily_start + 1
    for ci, h in enumerate(['猫', '干粮(g)', '猫饭(g)', '总热量(kcal)'], 1):
        ws.cell(row=dh, column=ci, value=h)
    style_header(ws, dh, 1, 4)
    density_ref = f'B{density_row}'
    cats = [('公猫A (5.8kg)', 30, 60), ('母猫B (4.5kg)', 30, 50), ('幼猫C (2.5kg)', 40, 80)]
    for ci, (cname, kibble, wet) in enumerate(cats):
        row = dh + 1 + ci
        style_cell(ws, row, 1, align=LEFT).value = cname
        style_cell(ws, row, 2, fmt='0').value = kibble
        style_cell(ws, row, 3, fmt='0').value = wet
        style_cell(ws, row, 4, fmt='0.0').value = f'={kibble}*4+{wet}*{density_ref}'

    total_daily_row = dh + 4 + 1
    style_cell(ws, total_daily_row, 1, align=LEFT).value = '合计'
    style_cell(ws, total_daily_row, 1).font = Font(name='微软雅黑', bold=True)
    style_cell(ws, total_daily_row, 4, fmt='0.0').value = f'=SUM(D{dh+2}:D{dh+4})'

    # 冻结窗格
    ws.freeze_panes = 'A7'

    return ws


# ── 主程序 ────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    for ver_name, ver_data in VERSIONS.items():
        create_version_sheet(wb, ver_name, ver_data)

    create_matrix_sheet(wb)

    out = r'c:\Users\nice-chocolate\Desktop\cat\猫饭NRC配方计算器.xlsx'
    wb.save(out)
    print(f'✅ 已生成: {out}')
    print(f'   包含工作表: {", ".join(wb.sheetnames)}')

if __name__ == '__main__':
    main()
